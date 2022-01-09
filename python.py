from openpyxl import load_workbook
wb = load_workbook('employeedata.xlsx')
page = wb.active

for i in range(2, page.max_row+1):
    cell = page.cell(i, 2)
   
    if 'helpinghands.cm' in cell.value:
        update = (cell.value).replace('helpinghand.cm','handsinhands.org')
        page.cell(i,2).value = update

wb.save('updated_employeedata.csv')
wb.save('updated_employeedata.xlsx')